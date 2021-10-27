import openpyxl

filePath = 'E:\ChenHao\Python\Project\AsianTeams.xlsx'
wb = openpyxl.load_workbook(filePath)
for n in range(0, len(wb.sheetnames) - 1):
    print(wb.sheetnames[n])
    sheet = wb[wb.sheetnames[n]]
    # sheet = wb.active 上一行已经按名称激活了sheet，所以无需再次激活。
    li2 = []
    for i in range(1, sheet.max_row):
        li = []
        for j in range(1, sheet.max_column + 1):
            li.append(sheet.cell(i, j).value)
        li2.append(li)
    print(li2)
